#!/usr/bin/env python3
"""Excel to JSON converter script."""

import json
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


def convert_value(cell):
    """Convert cell value to JSON-compatible type."""
    value = cell.value
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def excel_to_json(input_path, sheet=None, output_path=None, compact=False):
    """Convert Excel file to JSON."""
    wb = load_workbook(input_path, data_only=True)
    
    # Select sheet
    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    
    # Get all rows
    rows = list(ws.iter_rows())
    if not rows:
        return []
    
    # First row as headers
    headers = [cell.value for cell in rows[0]]
    
    # Convert remaining rows to objects
    data = []
    for row in rows[1:]:
        obj = {}
        for header, cell in zip(headers, row):
            if header is not None:
                obj[header] = convert_value(cell)
        data.append(obj)
    
    # Determine output path
    if output_path is None:
        output_path = Path(input_path).stem + ".json"
    
    # Write JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        if compact:
            json.dump(data, f, ensure_ascii=False)
        else:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 转换完成: {output_path}")
    print(f"  共 {len(data)} 条记录")
    return data


def main(params):
    """API 入口函数"""
    files = params.get('files', [])
    file_path = params.get('file_path', '')

    input_file = files[0] if files else file_path
    if not input_file:
        return {'status': 'error', 'message': '请上传 Excel 文件'}

    try:
        data = excel_to_json(input_file)

        # 生成输出文件
        output_path = OUTPUTS_DIR / generate_unique_filename('output', 'json')
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return {
            'status': 'success',
            'message': f'转换完成，共 {len(data)} 条记录',
            'content': data,
            '_output_file': {
                'path': str(output_path),
                'type': 'json',
                'name': output_path.name,
                'url': f'/outputs/{output_path.name}'
            }
        }
    except Exception as e:
        return {'status': 'error', 'message': str(e)}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Excel to JSON")
    parser.add_argument("input", help="Input Excel file path")
    parser.add_argument("--sheet", help="Sheet name or index (0-based)")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--compact", action="store_true", help="Output compact JSON")

    args = parser.parse_args()

    # Parse sheet argument
    sheet = args.sheet
    if sheet is not None and sheet.isdigit():
        sheet = int(sheet)

    excel_to_json(args.input, sheet, args.output, args.compact)